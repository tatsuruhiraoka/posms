# posms/cli.py
"""
posms.cli
=========

Command-line interface for the Postal Operation Shift-Management System.

Usage examples
--------------

# 予測 + シフト最適化 (E2E)
poetry run posms run-monthly --predict-date 2025-08-01

# 予測モデルだけ再学習（normal + registered_plus）
poetry run posms train --office-id 1

# 指定開始日から4週間の予測をDBへ書き戻し（mail_kind 指定）
poetry run posms forecast --office-id 1 --start 2025-09-08 --days 28 --mail-kind normal
poetry run posms forecast --office-id 1 --start 2025-09-08 --days 28 --mail-kind registered_plus

# シフトだけ再最適化（分担表案）
poetry run posms optimize --date 2025-08-01 --output-type 分担表案

# 初期マスタ投入（任意：PostgreSQL環境の初期化に便利）
poetry run posms seed-masters
"""

from __future__ import annotations

# --- standard library ---
import json
import os
import re
from datetime import date
from enum import Enum
from pathlib import Path
from shutil import copyfile
from typing import Annotated

# --- third-party ---
import pandas as pd
import typer
from sqlalchemy import inspect, text

# --- local application ---
from posms.utils.db import SessionManager

app = typer.Typer(help="Postal Operation Shift-Management System CLI")

class OutputType(str, Enum):  # Typer が扱える
    分担表 = "分担表"
    勤務指定表 = "勤務指定表"
    分担表案 = "分担表案"
    
# ---------- Helper -------------------------------------------------
def _default_template() -> Path:
    return Path("excel_templates/shift_template.xlsx")


def _engine():
    """SessionManager 経由で Engine を取得（Postgres⇄SQLite 自動切替）"""
    return SessionManager().engine


def _has_table(eng, name: str) -> bool:
    """方言差を吸収してテーブル存在を確認"""
    insp = inspect(eng)
    schema = "public" if eng.dialect.name == "postgresql" else None
    try:
        return insp.has_table(name, schema=schema)
    except TypeError:
        return insp.has_table(name)


def _strip_quotes(name: str) -> str:
    s = str(name).strip()
    if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
        return s[1:-1]
    return s


def _has_column(eng, table_name: str, column_name: str) -> bool:
    """テーブルのカラム有無を方言差も含めて確認（ケース差も吸収）"""
    insp = inspect(eng)
    schema = "public" if eng.dialect.name == "postgresql" else None
    t = _strip_quotes(table_name)

    try:
        cols = insp.get_columns(t, schema=schema)
    except TypeError:
        cols = insp.get_columns(t)
    except Exception:
        try:
            cols = insp.get_columns(t)
        except Exception:
            cols = []

    names = set()
    for c in cols:
        if isinstance(c, dict) and "name" in c:
            names.add(str(c["name"]).lower())
        else:
            try:
                names.add(str(c.name).lower())
            except Exception:
                pass
    return column_name.lower() in names

def _prepare_out_from_template(out: Path, template: Path) -> None:
    """
    out が無ければ、template (xlsm) をそのままコピーして作る。
    既に out があれば何もしない（テンプレの分担予定表(案)やマクロを保持）。
    """
    out = Path(out)
    template = Path(template)
    out.parent.mkdir(parents=True, exist_ok=True)
    if not out.exists():
        if not template.exists():
            raise FileNotFoundError(f"Template not found: {template}")
        copyfile(template, out)


def _truthy_int(x) -> int:
    """'○', '◯', 'true', '1', 'yes' などを 1、'×', 'false', '0', 'no' を 0 に正規化（値は厳格化対象外）"""
    s = str(x).strip().lower()
    if s in ("1", "true", "t", "yes", "y", "on", "○", "◯"):
        return 1
    if s in ("0", "false", "f", "no", "n", "off", "×"):
        return 0
    try:
        return 1 if int(float(s)) != 0 else 0
    except Exception:
        return 0


def _extract_first_int(x: object, default: int = 10**9) -> int:
    m = re.search(r"(\d+)", str(x))
    return int(m.group(1)) if m else default


def _sort_by_display_order_or_natural(
    df: pd.DataFrame,
    code_col: str,
    display_col: str = "__display_order__",
) -> pd.DataFrame:
    """
    display_order が有効ならそれで昇順。
    無ければ code_col（例: 'Z10'）を自然順（数値）で昇順。
    """
    out = df.copy()

    if (
        display_col in out.columns
        and pd.to_numeric(out[display_col], errors="coerce").notna().any()
    ):
        out["__ord__"] = (
            pd.to_numeric(out[display_col], errors="coerce").fillna(10**9).astype(int)
        )
        sort_cols = ["__ord__"]
        if code_col in out.columns:
            sort_cols.append(code_col)
        out = out.sort_values(sort_cols, kind="mergesort")
        out = out.drop(columns=["__ord__"])
        return out

    if code_col in out.columns:
        out["__ord__"] = out[code_col].astype(str).map(_extract_first_int).astype(int)
        out = out.sort_values(["__ord__", code_col], kind="mergesort").drop(
            columns=["__ord__"]
        )
        return out

    return out


def _write_df_overwrite_sheet(
    df: pd.DataFrame,
    out_path: Path,
    sheet_name: str,
    template_path: Path | None = None,
    start_cell: str = "A1",
    *,
    clear_to_df_area_only: bool = True,
) -> None:
    """
    Excelに「二重出力」させないための上書き書き込み（高速・安全版）。

    改善点:
    - 既存 ws.max_row / ws.max_column まで消しに行かない（テンプレが大きいと激遅になるため）
    - df に必要な範囲だけをクリアして書き込む
    - xlsm は keep_vba=True でマクロ保持

    clear_to_df_area_only=True:
      クリア範囲 = dfの書き込みに必要な領域（推奨）
    clear_to_df_area_only=False:
      旧挙動に近く、ws.max_row/max_col まで広げてクリア（遅い可能性あり）
    """
    import openpyxl
    from openpyxl.utils.cell import coordinate_to_tuple

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    tpl = Path(template_path) if template_path else None

    load_path: Path | None = None
    if tpl is not None and tpl.exists():
        load_path = tpl
    elif out_path.exists():
        load_path = out_path

    if load_path is None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        keep_vba = False
    else:
        keep_vba = (load_path.suffix.lower() == ".xlsm")
        wb = openpyxl.load_workbook(load_path, keep_vba=keep_vba)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    start_row, start_col = coordinate_to_tuple(start_cell)

    # --- df サイズ（ヘッダ行 + データ行）
    nrows = int(len(df.index))
    ncols = int(len(df.columns))
    if ncols <= 0:
        # 列が無いdfは何もしない（シートは作る/存在させる）
        wb.save(out_path)
        wb.close()
        return

    # 書き込み最終セル（ヘッダ1行分を含む）
    last_row_needed = start_row + nrows  # header=start_row, data starts start_row+1, ends start_row+nrows
    last_col_needed = start_col + (ncols - 1)

    # --- クリア範囲の決定
    if clear_to_df_area_only:
        clear_last_row = last_row_needed
        clear_last_col = last_col_needed
    else:
        clear_last_row = max(ws.max_row, last_row_needed)
        clear_last_col = max(ws.max_column, last_col_needed)

    # --- 値クリア（スタイルは残す）
    # 必要範囲だけ消すので高速
    for r in range(start_row, clear_last_row + 1):
        row_cells = ws.iter_rows(
            min_row=r, max_row=r, min_col=start_col, max_col=clear_last_col
        )
        for cells in row_cells:
            for cell in cells:
                cell.value = None

    # --- ヘッダ
    for j, col_name in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j).value = str(col_name)

    # --- データ（NaN/NaT → None）
    # itertuples は to_numpy より dtype 崩れにくく、行処理に向く
    for i, row in enumerate(df.itertuples(index=False, name=None), start=start_row + 1):
        for j, v in enumerate(row, start=start_col):
            if v is None or (isinstance(v, float) and pd.isna(v)) or pd.isna(v):
                v = None
            ws.cell(row=i, column=j).value = v

    wb.save(out_path)
    wb.close()
    
def _export_model_bundle_from_run(
    *,
    run_id: str,
    kind: str,
    feature_columns: list[str],
    out_dir: Path = Path("model_bundle"),
    experiment: str | None = None,
) -> Path:
    """
    MLflow の run から model/model.xgb を取り出して model_bundle/<kind>/ を作る。
    """
    import mlflow
    from mlflow.tracking import MlflowClient
    tracking_uri = mlflow.get_tracking_uri()
    client = MlflowClient(tracking_uri=tracking_uri)

    dst = out_dir / kind
    dst.mkdir(parents=True, exist_ok=True)

    # model.xgb を取得（あなたの保存形式に合わせる：artifact_path="model" / file="model.xgb"）
    local = client.download_artifacts(run_id, "model/model.xgb", dst.as_posix())
    src = Path(local)
    (dst / "model.xgb").write_bytes(src.read_bytes())

    # 特徴量列（順序が大事）
    (dst / "features.json").write_text(
        json.dumps(list(feature_columns), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    meta = {
        "run_id": run_id,
        "mail_kind": kind,
        "experiment": experiment,
        "tracking_uri": tracking_uri,
    }
    (dst / "meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return dst


# ---------- Commands ----------------------------------------------
@app.command("run-monthly")
def run_monthly(
    predict_date: Annotated[
        str,
        typer.Option("--predict-date", "-d", help="YYYY-MM-DD 形式。省略時は今日。"),
    ] = "",
    output_type: Annotated[
        OutputType,
        typer.Option(
            "--output-type",
            "-o",
            case_sensitive=False,
            help="Excel 出力の種類: 分担表 / 勤務指定表 / 分担表案",
        ),
    ] = OutputType.分担表,
    excel_template: Annotated[
        Path, typer.Option("--template", "-t", help="Excel テンプレート .xlsx")
    ] = _default_template(),
):
    """
    1. Excel → DB 取込
    2. XGBoost 再学習 (MLflow ログ)
    3. 需要予測 → PuLP シフト最適化
    4. Excel に書き出し
    """
    params = {
        "predict_date": predict_date or str(date.today()),
        "output_type": output_type.value,
        "excel_template": str(excel_template),
    }
    from posms.flows import monthly_flow as mf

    _monthly = getattr(mf, "monthly_refresh", None) or getattr(mf, "monthly_train")
    _monthly(**params)


@app.command("train")
def train_model(
    office_id: Annotated[
        int | None, typer.Option("--office-id", help="局ID（1局のみなら省略可）")
    ] = None,
    n_estimators: Annotated[int, typer.Option(help="XGBoost 木数")] = 200,
    max_depth: Annotated[int, typer.Option(help="XGBoost 木の深さ")] = 6,
):
    """DB を使ってモデルを学習し、MLflowに保存。normal + registered_plus を両方実行。"""
    import os
    from pathlib import Path

    import pandas as pd
    import typer
    from mlflow.tracking import MlflowClient

    # 遅延 import（export系の import 連鎖を防ぐ）
    from posms.features.builder import FeatureBuilder
    from posms.models.normal.trainer import ModelTrainer as NormalTrainer
    from posms.models.registered_plus.pipeline import train_from_hist as rp_train_from_hist
    from posms.models.registered_plus.features import FEATURES_REGISTERED_PLUS

    # ====== 重要：モデル名を分離（事故防止） ======
    MODEL_NAME_NORMAL = "posms_normal"
    MODEL_NAME_REGPLUS = "posms_registered_plus"
    experiment = os.getenv("MLFLOW_EXPERIMENT_NAME", "posms")

    params = {"n_estimators": n_estimators, "max_depth": max_depth}

    # -------------------------
    # normal
    # -------------------------
    fb_n = FeatureBuilder(office_id=office_id, mail_kind="normal")
    Xn, yn = fb_n.build()

    tags_normal = {
        "model_name": MODEL_NAME_NORMAL,
        "mail_kind": "normal",
        "office_id": str(fb_n.office_id),
        "feature_set": ",".join(map(str, list(Xn.columns))),
    }

    run_id_normal = NormalTrainer(params=params, experiment=experiment).train(
        Xn,
        yn,
        auto_register=False,
        tags=tags_normal,
    )

    # -------------------------
    # registered_plus（registered + lp_plus の sum）
    # -------------------------
    fb_reg = FeatureBuilder(office_id=fb_n.office_id, mail_kind="registered", engine=fb_n.engine)
    fb_lp  = FeatureBuilder(office_id=fb_n.office_id, mail_kind="lp_plus",    engine=fb_n.engine)
    
    reg = fb_reg._load_mail()[["date", "actual_volume"]].rename(columns={"actual_volume": "registered"})
    lp  = fb_lp._load_mail()[["date", "actual_volume"]].rename(columns={"actual_volume": "lp_plus"})
    
    m = reg.merge(lp, on="date", how="outer").sort_values("date").fillna(0)
    m["sum"] = m["registered"] + m["lp_plus"]
    
    # 学習用（index=DatetimeIndex、sumを含む）
    df_hist = m.set_index(pd.to_datetime(m["date"]))[["registered", "lp_plus", "sum"]]

    tags_rp = {
        "model_name": MODEL_NAME_REGPLUS,
        "mail_kind": "registered_plus",
        "office_id": str(fb_n.office_id),
        "feature_set": ",".join(map(str, list(FEATURES_REGISTERED_PLUS))),
    }

    res = rp_train_from_hist(
        df_hist=df_hist,
        experiment_name=experiment,
        run_name=f"{MODEL_NAME_REGPLUS}-office{fb_n.office_id}",
        tags=tags_rp,
        model_name=MODEL_NAME_REGPLUS,
    )
    run_id_regplus = res.run_id if hasattr(res, "run_id") else res

    # ===== model_bundle 自動生成 =====
    out_root = Path("model_bundle")

    bundle_n = _export_model_bundle_from_run(
        run_id=run_id_normal,
        kind="normal",
        feature_columns=list(Xn.columns),
        out_dir=out_root,
        experiment=experiment,
    )

    bundle_rp = _export_model_bundle_from_run(
        run_id=run_id_regplus,
        kind="registered_plus",
        feature_columns=list(FEATURES_REGISTERED_PLUS),
        out_dir=out_root,
        experiment=experiment,
    )

    # ===== タグ存在チェック（必ず刻む）=====
    client = MlflowClient()
    for rid in [run_id_normal, run_id_regplus]:
        run_tags = client.get_run(rid).data.tags
        for k in ["model_name", "mail_kind", "office_id", "feature_set"]:
            if k not in run_tags:
                raise RuntimeError(f"missing tag {k} in run {rid}")

    typer.echo(f"✅ model_bundle(normal): {bundle_n}")
    typer.echo(f"✅ model_bundle(registered_plus): {bundle_rp}")

@app.command("forecast")
def forecast_4weeks(
    start: Annotated[str, typer.Option("--start", "-s", help="YYYY-MM-DD")],
    days: Annotated[int, typer.Option("--days", "-n")] = 28,
    office_id: Annotated[int | None, typer.Option("--office-id")] = None,
    mail_kind: Annotated[str, typer.Option("--mail-kind")] = "normal",
):
    import sys
    import numpy as np
    import pandas as pd
    import jpholiday
    import xgboost as xgb
    from sqlalchemy import text

    from posms.features.builder import FeatureBuilder, FEATURE_COLUMNS
    from posms.models.predictor import ModelPredictor

    # ------------------------------------------------------------
    # SQLite DATABASE_URL 自動設定（空DB事故防止）
    # ------------------------------------------------------------
    if not os.getenv("DATABASE_URL"):
        p = os.getenv("POSMS_DB_PATH")
        if p:
            db_path = Path(p).expanduser().resolve()
        else:
            base = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path.cwd()
            db_path = (base / "posms.db").resolve()

        if not db_path.exists():
            raise RuntimeError(
                f"DB connection info is incomplete.\n"
                f"SQLite file not found: {db_path}\n"
                f"Set DATABASE_URL or POSMS_DB_PATH."
            )

        os.environ["DATABASE_URL"] = f"sqlite:///{db_path.as_posix()}"

    if mail_kind not in ("normal", "registered_plus"):
        raise typer.BadParameter(mail_kind)

    start_ts = pd.to_datetime(start).normalize()
    end_ts = start_ts + pd.Timedelta(days=days - 1)

    # ------------------------------------------------------------
    # DB engine / office_id 解決（FeatureBuilder の仕様に従う）
    # ------------------------------------------------------------
    fb_base = FeatureBuilder(office_id=office_id, mail_kind="normal")
    eng = fb_base.engine

    # office_id が None のままなら、normal の _load_mail が自動選択 or 例外
    _ = fb_base._load_mail()
    office_id = int(fb_base.office_id)

    # ------------------------------------------------------------
    # DB 行保証
    # ------------------------------------------------------------
    def ensure_rows(kind: str, d1: pd.Timestamp, d2: pd.Timestamp):
        with eng.begin() as con:
            for d in pd.date_range(d1, d2, freq="D"):
                con.execute(
                    text(
                        """
                        INSERT INTO mailvolume_by_type
                          ("date", office_id, mail_kind, actual_volume, forecast_volume)
                        VALUES
                          (:d, :o, :k, NULL, NULL)
                        ON CONFLICT ("date", office_id, mail_kind) DO NOTHING
                        """
                    ),
                    {"d": d.date(), "o": office_id, "k": kind},
                )

    # ==========================================================
    # normal（rolling + posms_normal）
    # ==========================================================
    if mail_kind == "normal":
        from posms.models.rolling import rolling_forecast_28d
    
        # FeatureBuilder は特徴量生成だけ
        fb = FeatureBuilder(office_id=office_id, mail_kind="normal", engine=eng)
    
        # 予測器
        pred = ModelPredictor(
            model_name="posms_normal",
            mail_kind="normal",
            office_id=int(office_id),
            experiment=os.getenv("MLFLOW_EXPERIMENT_NAME", "posms"),
            tracking_uri=os.getenv("MLFLOW_TRACKING_URI"),
            stage=None,
        )
    
        # rolling
        res = rolling_forecast_28d(
            fb=fb,
            predictor=pred,
            start=start_ts.date(),
            days=days,
            context_days=7,
        )
    
        # start〜end の範囲だけ切り出し（date型を正規化してから比較）
        df_out = res.df.copy()
        df_out["date"] = pd.to_datetime(df_out["date"]).dt.normalize()
        df_out = df_out[(df_out["date"] >= start_ts) & (df_out["date"] <= end_ts)].copy()
    
        if df_out.empty:
            typer.echo("更新対象なし（rolling出力が空）")
            return
    
        # raw（日次の発生量予測）: y_pred優先、NaNはy_filledで補完
        if "y_pred" not in df_out.columns or "y_filled" not in df_out.columns:
            raise RuntimeError(f"rolling result missing y_pred/y_filled. cols={list(df_out.columns)}")
    
        raw = df_out.set_index("date")["y_pred"].astype(float)
        raw = raw.fillna(df_out.set_index("date")["y_filled"].astype(float))
        raw = raw.asfreq("D")
    
        # ------------------------------------------------------------
        # 丸め + 土日祝繰越（配達量へ変換）
        # ------------------------------------------------------------
        df_deliver = ModelPredictor.apply_delivery_rules(
            raw,
            round_to_thousand=True,
            extend_to_next_delivery=True,
        )
    
        df_deliver["date"] = pd.to_datetime(df_deliver["date"]).dt.normalize()
        df_deliver = df_deliver[(df_deliver["date"] >= start_ts) & (df_deliver["date"] <= end_ts)].copy()
    
        updates = [(d.date(), int(v)) for d, v in zip(df_deliver["date"], df_deliver["deliver_pred"])]
    
        if not updates:
            typer.echo("更新対象なし（予測結果が空）")
            return
    
        # 行保証（UPDATE対象が無い事故防止）
        ensure_rows("normal", pd.Timestamp(updates[0][0]), pd.Timestamp(updates[-1][0]))
    
        # 書き戻し
        with eng.begin() as con:
            for d, v in updates:
                con.execute(
                    text(
                        """
                        UPDATE mailvolume_by_type
                           SET forecast_volume=:v
                         WHERE "date"=:d AND office_id=:o AND mail_kind='normal'
                        """
                    ),
                    {"v": v, "d": d, "o": int(office_id)},
                )
    
        typer.echo(
            f"forecast_volume 更新: {len(updates)} 件 "
            f"(mail_kind=normal, office_id={office_id}, {start_ts.date()}〜{end_ts.date()})"
        )
        return


    # ==========================================================
    # registered_plus（繰越なし：raw をそのまま DB に書く）
    # ※ DB の registered + lp_plus を合算して自己回帰ローリング
    # ==========================================================
    from posms.models.registered_plus.features import (
        FEATURES_REGISTERED_PLUS,
        build_registered_plus_feature_row,
    )
    from posms.models.predictor import ModelPredictor
    
    # 予測器（あなたの標準呼び出し）
    pred = ModelPredictor(
        model_name="posms_registered_plus",
        mail_kind="registered_plus",
        office_id=int(office_id),
        experiment=os.getenv("MLFLOW_EXPERIMENT_NAME", "posms"),
        tracking_uri=os.getenv("MLFLOW_TRACKING_URI"),
        stage=None,
    )
    
    # DBから実績を取る（mail_kind: registered / lp_plus）
    def _load_kind(kind: str) -> pd.Series:
        sql = """
            SELECT "date", actual_volume
              FROM mailvolume_by_type
             WHERE office_id=:o AND mail_kind=:k
             ORDER BY "date"
        """
        df = pd.read_sql(text(sql), eng, params={"o": office_id, "k": kind}, parse_dates=["date"])
        if df.empty:
            raise RuntimeError(f"registered_plus: no rows for mail_kind={kind!r}")
        s = df.set_index("date")["actual_volume"].astype(float)
        s.index = pd.to_datetime(s.index).normalize()
        return s
    
    s_reg = _load_kind("registered")
    s_lp = _load_kind("lp_plus")
    
    # 合算系列（過去実績）
    vol = s_reg.add(s_lp, fill_value=0.0).sort_index()
    
    last_actual_date = vol.dropna().index.max()
    if last_actual_date is None:
        raise RuntimeError("registered_plus: base series is empty")
    
    # 実績翌日から予測で埋める（start が先ならギャップも内部で埋める）
    bridge_start = pd.to_datetime(last_actual_date).normalize() + pd.Timedelta(days=1)
    start_for_pred = bridge_start if start_ts > bridge_start else start_ts
    full_dates = pd.date_range(start_for_pred, end_ts, freq="D")
    
    # 連続化（未来まで index を用意）
    full_idx = pd.date_range(vol.index.min(), end_ts, freq="D")
    vol = vol.reindex(full_idx).fillna(0.0)
    
    updates: dict[pd.Timestamp, float] = {}
    
    for dt in full_dates:
        feat_dict = build_registered_plus_feature_row(dt, vol)
    
        # 列順は FEATURES_REGISTERED_PLUS（唯一の真実）
        X = pd.DataFrame([feat_dict])[list(FEATURES_REGISTERED_PLUS)].astype(float)
    
        yhat = float(pred.predict(X)[0])
        yhat = max(0.0, yhat)
    
        # 次の日の lag 用に vol を更新（registered_plus は繰越しない）
        vol.loc[pd.to_datetime(dt).normalize()] = yhat
        updates[pd.to_datetime(dt).normalize()] = yhat
    
    s_pred = pd.Series(updates).sort_index()
    s_pred = s_pred[(s_pred.index >= start_ts) & (s_pred.index <= end_ts)]
    
    if s_pred.empty:
        typer.echo("更新対象なし（直近実績不足 or 特徴量不足）")
        return
    
    df_w = s_pred.rename("forecast_volume").reset_index().rename(columns={"index": "date"})
    df_w["date"] = pd.to_datetime(df_w["date"]).dt.normalize()
    df_w["forecast_volume"] = df_w["forecast_volume"].astype(float).clip(lower=0.0).round().astype(int)
    
    ensure_rows("registered_plus", df_w["date"].min(), df_w["date"].max())
    
    with eng.begin() as con:
        for _, r in df_w.iterrows():
            con.execute(
                text(
                    """
                    UPDATE mailvolume_by_type
                       SET forecast_volume=:v
                     WHERE "date"=:d AND office_id=:o AND mail_kind='registered_plus'
                    """
                ),
                {"v": int(r["forecast_volume"]), "d": pd.to_datetime(r["date"]).date(), "o": int(office_id)},
            )
    
    typer.echo(
        f"forecast_volume 更新: {len(df_w)} 件 "
        f"(mail_kind=registered_plus, office_id={office_id}, {start_ts.date()}〜{end_ts.date()})"
    )
    return

    

@app.command("export-sheet-employees")
def export_sheet_employees(
    department_code: str = typer.Option(
        ..., "--department-code", "-dc", help="例: DPT-A もしくは 部署名"
    ),
    team: str = typer.Option(..., "--team", "-t", help="例: 1班"),
    out: Path = typer.Option(Path("excel_out/班データ.xlsx"), "--out"),
    sheet: str = typer.Option("社員", "--sheet"),
    template: Path = typer.Option(None, "--template"),
):
    eng = _engine()

    insp = inspect(eng)
    dept_cols = {
        c["name"] if isinstance(c, dict) else c.name
        for c in insp.get_columns("department")
    }
    has_code = "department_code" in dept_cols

    base_sql = """
      SELECT
          e.employee_code   AS "社員番号",
          e.name            AS "氏名",
          d.department_name AS "部",
          t.team_name       AS "班",
          COALESCE(e.employment_type,'')  AS "社員タイプ",
          COALESCE(e.position,'')         AS "役職",
          CASE WHEN COALESCE(e.is_leader,0)      <> 0 THEN 1 ELSE 0 END AS "班長",
          CASE WHEN COALESCE(e.is_vice_leader,0) <> 0 THEN 1 ELSE 0 END AS "副班長",
          CASE WHEN COALESCE(e.is_certifier,0)   <> 0 THEN 1 ELSE 0 END AS "認証司",
          COALESCE(e.default_work_hours,0)  AS "勤務時間(日)",
          COALESCE(e.monthly_work_hours,0)  AS "勤務時間(月)"
      FROM employee e
      JOIN team t ON t.team_id = e.team_id
      JOIN department d ON d.department_id = t.department_id
      WHERE {where_clause}
    """

    if has_code:
        where_clause = "(COALESCE(d.department_code,'') = :dc OR d.department_name = :dc) AND t.team_name = :tn"
    else:
        where_clause = "d.department_name = :dc AND t.team_name = :tn"

    sql = base_sql.format(where_clause=where_clause)

    with eng.connect() as con:
        df = pd.read_sql(text(sql), con, params={"dc": department_code, "tn": team})

    # 自然順ソート：社員番号の数字部分で並び替え
    if "社員番号" in df.columns:
        df["__num__"] = (
            df["社員番号"]
            .astype(str)
            .str.extract(r"(\d+)", expand=False)
            .fillna("999999999")
            .astype(int)
        )
        df = df.sort_values(["__num__", "社員番号"], kind="mergesort").drop(
            columns="__num__"
        )

    for c in ["班長", "副班長", "認証司", "勤務時間(日)", "勤務時間(月)"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)

    _write_df_overwrite_sheet(
        df=df,
        out_path=out,
        sheet_name=sheet,
        template_path=template,
        start_cell="A1",
    )
    typer.echo(f"✅ 社員シート {len(df)}行 → {out}（{sheet}）")


@app.command("export-sheet-zones")
def export_sheet_zones(
    department_code: str = typer.Option(..., "--department-code", "-dc"),
    team: str = typer.Option(..., "--team", "-t"),
    out: Path = typer.Option(Path("excel_out/班データ.xlsx"), "--out"),
    sheet: str = typer.Option("区情報", "--sheet"),
    template: Path = typer.Option(None, "--template"),
):
    eng = _engine()

    insp = inspect(eng)
    dept_cols = {
        c["name"] if isinstance(c, dict) else c.name
        for c in insp.get_columns("department")
    }
    has_code = "department_code" in dept_cols

    has_zone_order = _has_column(eng, "zone", "display_order")
    order_select = (
        'z.display_order AS "__display_order__"'
        if has_zone_order
        else 'NULL AS "__display_order__"'
    )

    base_sql = f"""
      SELECT
        CAST(COALESCE(z.zone_code, 'Z' || z.zone_id) AS TEXT) AS "区コード",
        COALESCE(z.zone_name,'')     AS "区名",
        t.team_name                  AS "班",
        COALESCE(z.operational_status,'通配') AS "稼働",
        COALESCE(dem.demand_mon,0)     AS "月",
        COALESCE(dem.demand_tue,0)     AS "火",
        COALESCE(dem.demand_wed,0)     AS "水",
        COALESCE(dem.demand_thu,0)     AS "木",
        COALESCE(dem.demand_fri,0)     AS "金",
        COALESCE(dem.demand_sat,0)     AS "土",
        COALESCE(dem.demand_sun,0)     AS "日",
        COALESCE(dem.demand_holiday,0) AS "祝",
        COALESCE(z.shift_type,'日勤')   AS "シフトタイプ",
        {order_select}
      FROM zone z
      JOIN team t          ON t.team_id = z.team_id
      JOIN department d    ON d.department_id = t.department_id
      LEFT JOIN demandprofile dem ON dem.zone_id = z.zone_id
      WHERE {{where_clause}}
    """

    if has_code:
        where_clause = "(COALESCE(d.department_code,'') = :dc OR d.department_name = :dc) AND t.team_name = :tn"
    else:
        where_clause = "d.department_name = :dc AND t.team_name = :tn"

    sql = base_sql.format(where_clause=where_clause)

    with eng.connect() as con:
        df = pd.read_sql(text(sql), con, params={"dc": department_code, "tn": team})

    # display_order優先、無ければ区コード自然順
    df = _sort_by_display_order_or_natural(
        df, code_col="区コード", display_col="__display_order__"
    )
    if "__display_order__" in df.columns:
        df = df.drop(columns=["__display_order__"])

    _write_df_overwrite_sheet(
        df=df,
        out_path=out,
        sheet_name=sheet,
        template_path=template,
        start_cell="A1",
    )
    typer.echo(f"✅ 区情報シート {len(df)}行 → {out}（{sheet}）")


@app.command("export-sheet-employee-demand")
def export_sheet_employee_demand(
    team: str = typer.Option(..., "--team", "-t", help="例: 1班"),
    out: Path = typer.Option(Path("excel_out/班データ.xlsx"), "--out"),
    template: Path = typer.Option(
        None, "--template", help="xls/xlsx/xlsm（xlsmはマクロ温存）"
    ),
    sheet: str = typer.Option("社員別需要", "--sheet"),
):
    """
    班ごとの『社員別需要』シートをExcelに出力:
      列 = 社員番号, 氏名, <区名...横展開>, 月, 火, 水, 木, 金, 土, 日, 祝
      値 = 区列: EmployeeZoneProficiency.proficiency（無い所は0） / 曜日列: 1/0
    """
    eng = _engine()

    has_zone_order = _has_column(eng, "zone", "display_order")
    zone_order_select = (
        "z.display_order AS display_order"
        if has_zone_order
        else "NULL AS display_order"
    )

    sql_zones = f"""
      SELECT z.zone_id, z.zone_code, z.zone_name, {zone_order_select}
        FROM zone z
        JOIN team t ON z.team_id = t.team_id
       WHERE t.team_name = :team
    """

    sql_emp = """
      SELECT e.employee_id, e.employee_code, e.name,
             CASE WHEN COALESCE(ea.available_mon,  1) <> 0 THEN 1 ELSE 0 END AS mon,
             CASE WHEN COALESCE(ea.available_tue,  1) <> 0 THEN 1 ELSE 0 END AS tue,
             CASE WHEN COALESCE(ea.available_wed,  1) <> 0 THEN 1 ELSE 0 END AS wed,
             CASE WHEN COALESCE(ea.available_thu,  1) <> 0 THEN 1 ELSE 0 END AS thu,
             CASE WHEN COALESCE(ea.available_fri,  1) <> 0 THEN 1 ELSE 0 END AS fri,
             CASE WHEN COALESCE(ea.available_sat,  0) <> 0 THEN 1 ELSE 0 END AS sat,
             CASE WHEN COALESCE(ea.available_sun,  0) <> 0 THEN 1 ELSE 0 END AS sun,
             CASE WHEN COALESCE(ea.available_hol,  0) <> 0 THEN 1 ELSE 0 END AS hol,
             COALESCE(ea.available_early,     '') AS early,
             COALESCE(ea.available_day,       '') AS day,
             COALESCE(ea.available_mid,       '') AS mid,
             COALESCE(ea.available_night,     '') AS night,
             COALESCE(ea.available_night_sat, '') AS night_sat,
             COALESCE(ea.available_night_sun, '') AS night_sun,
             COALESCE(ea.available_night_hol, '') AS night_hol
        FROM employee e
        JOIN team t ON e.team_id = t.team_id
   LEFT JOIN employee_availabilities ea ON ea.employee_id = e.employee_id
       WHERE t.team_name = :team
    """

    sql_prof = """
      SELECT ezp.employee_id, z.zone_name, COALESCE(ezp.proficiency, 0) AS proficiency
        FROM employeezoneproficiency ezp
        JOIN zone z ON z.zone_id = ezp.zone_id
        JOIN team t ON z.team_id = t.team_id
       WHERE t.team_name = :team
    """

    with eng.connect() as con:
        df_z = pd.read_sql(text(sql_zones), con, params={"team": team})
        df_e = pd.read_sql(text(sql_emp), con, params={"team": team})
        df_p = pd.read_sql(text(sql_prof), con, params={"team": team})

    # 列名小文字化
    for dfx in (df_z, df_e, df_p):
        dfx.columns = [c.lower() for c in dfx.columns]

    # 区の並びを display_order優先→無ければ zone_code自然順
    if (
        "display_order" in df_z.columns
        and pd.to_numeric(df_z["display_order"], errors="coerce").notna().any()
    ):
        df_z["__ord__"] = (
            pd.to_numeric(df_z["display_order"], errors="coerce")
            .fillna(10**9)
            .astype(int)
        )
        df_z = df_z.sort_values(["__ord__", "zone_code"], kind="mergesort").drop(
            columns=["__ord__"]
        )
    else:
        if "zone_code" in df_z.columns:
            df_z["__ord__"] = (
                df_z["zone_code"].astype(str).map(_extract_first_int).astype(int)
            )
            df_z = df_z.sort_values(["__ord__", "zone_code"], kind="mergesort").drop(
                columns=["__ord__"]
            )
        else:
            df_z = df_z.sort_values(["zone_id"], kind="mergesort")

    # 社員コードの自然順
    if "employee_code" in df_e.columns:
        df_e["__num__"] = (
            df_e["employee_code"]
            .astype(str)
            .str.extract(r"(\d+)", expand=False)
            .fillna("999999999")
            .astype(int)
        )
        df_e = df_e.sort_values(["__num__", "employee_code"], kind="mergesort").drop(
            columns="__num__"
        )

    zone_cols = df_z["zone_name"].dropna().astype(str).tolist()

    # 熟練度を横展開
    if df_p.empty:
        wide = pd.DataFrame(columns=["employee_id"] + zone_cols)
    else:
        wide = (
            df_p.pivot_table(
                index="employee_id",
                columns="zone_name",
                values="proficiency",
                fill_value=0,
                aggfunc="max",
            )
            .reset_index()
            .rename_axis(None, axis=1)
        )

    df = df_e.merge(wide, on="employee_id", how="left")

    for zc in zone_cols:
        if zc not in df.columns:
            df[zc] = 0
        df[zc] = pd.to_numeric(df[zc], errors="coerce").fillna(0).astype("int64")

    for c in ["mon", "tue", "wed", "thu", "fri", "sat", "sun", "hol"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype("int64")

    ordered = (
        ["employee_code", "name"]
        + zone_cols
        + [
            "mon",
            "tue",
            "wed",
            "thu",
            "fri",
            "sat",
            "sun",
            "hol",
            "early",
            "day",
            "mid",
            "night",
            "night_sat",
            "night_sun",
            "night_hol",
        ]
    )
    ordered = [c for c in ordered if c in df.columns]
    df = df[ordered].rename(
        columns={
            "employee_code": "社員番号",
            "name": "氏名",
            "mon": "月",
            "tue": "火",
            "wed": "水",
            "thu": "木",
            "fri": "金",
            "sat": "土",
            "sun": "日",
            "hol": "祝",
            "early": "早番",
            "day": "日勤",
            "mid": "中勤",
            "night": "夜勤",
            "night_sat": "夜勤(土)",
            "night_sun": "夜勤(日)",
            "night_hol": "夜勤(祝)",
        }
    )

    _write_df_overwrite_sheet(
        df=df,
        out_path=out,
        sheet_name=sheet,
        template_path=template,
        start_cell="A1",
    )
    typer.echo(f"✅ 社員別需要シート {len(df)}行 → {out}（{sheet}）")


@app.command("export-sheet-jobtypes-regular")
def export_sheet_jobtypes_regular(
    out: Path = typer.Option(Path("excel_out/班データ統合.xlsx"), "--out"),
    template: Path = typer.Option(None, "--template"),
    sheet: str = typer.Option("勤務種別(正社員)", "--sheet"),
):
    eng = _engine()
    JT = "jobtype"
    jt_for_inspect = JT

    has_disp = _has_column(eng, jt_for_inspect, "display_order")
    disp_select = (
        'display_order AS "__display_order__",'
        if has_disp
        else 'NULL AS "__display_order__",'
    )

    sql = f"""
      SELECT
             {disp_select}
             job_name   AS "勤務名",
             start_time AS "就労開始時間",
             end_time   AS "就労終了時間",
             work_hours AS "勤務時間"
        FROM {JT}
       WHERE lower(COALESCE(classification,'')) IN ('reg','regular','fulltime','正社員')
    """

    with eng.connect() as con:
        df = pd.read_sql(text(sql), con)

    df = _sort_by_display_order_or_natural(
        df, code_col="勤務名", display_col="__display_order__"
    )
    if "__display_order__" in df.columns:
        df = df.drop(columns=["__display_order__"])

    _write_df_overwrite_sheet(
        df=df,
        out_path=out,
        sheet_name=sheet,
        template_path=template,
        start_cell="A1",
    )
    typer.echo(f"✅ 勤務種別(正社員) {len(df)}行 → {out}（{sheet}）")


@app.command("export-sheet-jobtypes-fixedterm")
def export_sheet_jobtypes_fixedterm(
    out: Path = typer.Option(Path("excel_out/班データ統合.xlsx"), "--out"),
    template: Path = typer.Option(None, "--template"),
    sheet: str = typer.Option("勤務種別(期間雇用)", "--sheet"),
):
    eng = _engine()
    JT = "jobtype"
    jt_for_inspect = JT

    has_disp = _has_column(eng, jt_for_inspect, "display_order")
    disp_select = (
        'display_order AS "__display_order__",'
        if has_disp
        else 'NULL AS "__display_order__",'
    )

    sql = f"""
      SELECT
             {disp_select}
             job_name   AS "勤務名",
             start_time AS "就労開始時間",
             end_time   AS "就労終了時間",
             work_hours AS "勤務時間"
        FROM {JT}
       WHERE lower(COALESCE(classification,'')) IN ('contract','part-time','pt','temp','dispatch','期間雇用社員')
    """

    with eng.connect() as con:
        df = pd.read_sql(text(sql), con)

    df = _sort_by_display_order_or_natural(
        df, code_col="勤務名", display_col="__display_order__"
    )
    if "__display_order__" in df.columns:
        df = df.drop(columns=["__display_order__"])

    _write_df_overwrite_sheet(
        df=df,
        out_path=out,
        sheet_name=sheet,
        template_path=template,
        start_cell="A1",
    )
    typer.echo(f"✅ 勤務種別(期間雇用) {len(df)}行 → {out}（{sheet}）")


@app.command("export-sheet-leavetypes")
def export_sheet_leavetypes(
    out: Path = typer.Option(Path("excel_out/班データ統合.xlsx"), "--out"),
    template: Path = typer.Option(None, "--template"),
    sheet: str = typer.Option("休暇種類", "--sheet"),
):
    eng = _engine()
    sql = """
      SELECT
        leave_name     AS "休暇名",
        leave_code     AS "休暇コード",
        leave_category AS "休暇種類"
      FROM leavetype
      ORDER BY leave_category, leave_code, leave_name
    """
    with eng.connect() as con:
        df = pd.read_sql(text(sql), con)

    _write_df_overwrite_sheet(
        df=df,
        out_path=out,
        sheet_name=sheet,
        template_path=template,
        start_cell="A1",
    )
    typer.echo(f"✅ 休暇種類 {len(df)}行 → {out}（{sheet}）")


@app.command("export-sheet-special-attendance")
def export_sheet_special_attendance(
    out: Path = typer.Option(Path("excel_out/班データ統合.xlsx"), "--out"),
    template: Path = typer.Option(None, "--template"),
    sheet: str = typer.Option("特殊区分", "--sheet"),
):
    eng = _engine()
    sql = """
      SELECT
        attendance_name   AS "区分名",
        attendance_code   AS "区分コード",
        holiday_work_flag AS "休日勤務",
        is_active         AS "有効"
      FROM special_attendance_type
      WHERE COALESCE(is_active, 1) <> 0
      ORDER BY attendance_code, attendance_name
    """
    with eng.connect() as con:
        df = pd.read_sql(text(sql), con)

    _write_df_overwrite_sheet(
        df=df,
        out_path=out,
        sheet_name=sheet,
        template_path=template,
        start_cell="A1",
    )
    typer.echo(f"✅ 特殊区分 {len(df)}行 → {out}（{sheet}）")


@app.command("export-team-workbook")
def export_team_workbook(
    department_code: str = typer.Option(
        ..., "--department-code", "-dc", help="例: DPT-A"
    ),
    team: str = typer.Option(..., "--team", "-t", help="班名（例: 1班）"),
    out: Path = typer.Option(
        Path("excel_templates/班統合データ.xlsm"), "--out", help="出力先 .xlsm"
    ),
    template: Path = typer.Option(
        Path("excel_templates/分担予定表(案).xlsm"),
        "--template",
        help="テンプレート（xlsmを指定すればマクロ保持。分担予定表(案)が入っていること）",
    ),
    # ▼ 後方互換用
    db_url: Annotated[str | None, typer.Option("--db-url")] = None,
    sqlite: Annotated[
        Path | None, typer.Option("--sqlite", help="sqlite DB ファイルパス")
    ] = None,
    plan_sheet_name: Annotated[
        str, typer.Option("--plan-sheet-name", help="テンプレ内のシート名確認用")
    ] = "分担予定表(案)",
):
    """
    部署・班ごとの Excel ブックをテンプレ込みで統合出力。
    シート構成（テンプレ保持）：
      - 分担予定表(案)
      - 社員 / 区情報 / 社員別需要 / 正社員服務表 / 期間雇用社員服務表 / 休暇種類 / 特殊区分
    """
    if sqlite is not None:
        os.environ["DATABASE_URL"] = f"sqlite:///{Path(sqlite).resolve()}"
    elif db_url:
        os.environ["DATABASE_URL"] = db_url

    typer.echo(f"▶ 班「{team}」の統合Excelを作成中...")

    _prepare_out_from_template(out, template)
    typer.echo(f"  - テンプレ基盤: {template} → {out}")

    # （任意）テンプレ検証
    try:
        import openpyxl

        wb = openpyxl.load_workbook(out, keep_vba=True, read_only=True)
        if plan_sheet_name not in wb.sheetnames:
            typer.secho(
                f"WARNING: 出力ブックに '{plan_sheet_name}' シートが見つかりません。テンプレの確認をしてください。",
                err=True,
            )
        wb.close()
    except Exception:
        pass

    # ★重要：二重出力防止のため、各シートは「上書き」で出す
    export_sheet_employees(
        department_code=department_code, team=team, out=out, sheet="社員", template=out
    )
    export_sheet_zones(
        department_code=department_code,
        team=team,
        out=out,
        sheet="区情報",
        template=out,
    )
    export_sheet_employee_demand(team=team, out=out, template=out, sheet="社員別需要")
    export_sheet_jobtypes_regular(out=out, template=out, sheet="正社員服務表")
    export_sheet_jobtypes_fixedterm(out=out, template=out, sheet="期間雇用社員服務表")
    export_sheet_leavetypes(out=out, template=out, sheet="休暇種類")
    export_sheet_special_attendance(out=out, template=out, sheet="特殊区分")

    # 仕上げ：休暇種類・特殊区分シートを VeryHidden / シート順を整える
    try:
        import openpyxl

        wb = openpyxl.load_workbook(out, keep_vba=True)

        # sheet order（見たい順）
        desired = [
            "分担予定表(案)",
            "社員",
            "区情報",
            "社員別需要",
            "正社員服務表",
            "期間雇用社員服務表",
            "休暇種類",
            "特殊区分",
        ]
        ordered_sheets = [wb[n] for n in desired if n in wb.sheetnames]
        rest = [ws for ws in wb.worksheets if ws.title not in desired]
        # openpyxl の実用的手段（private属性アクセス）
        wb._sheets = ordered_sheets + rest  # noqa: SLF001

        for s in ("休暇種類", "特殊区分"):
            if s in wb.sheetnames:
                wb[s].sheet_state = "veryHidden"

        wb.save(out)
        wb.close()
    except Exception as e:
        typer.secho(f"NOTE: 後処理に失敗: {e}", err=True)

    typer.echo(f"✅ 統合完了: {out.resolve()}")


@app.command("import-excel")
def import_excel(
    file: Path = typer.Option(
        ..., "--file", exists=True, help="編集済みの班ファイル .xlsx"
    ),
):
    """
    Excel から DB へ取り込み（UPSERT）
    - C) 社員: employee（部/班マスタも補完）
    - B) 区情報: zone → demandprofile
    - A) 社員別需要: employee_availabilities/ employeezoneproficiency
    - D) 休暇種類: leavetype（シート名: 休暇種類 / 列: 休暇コード, 休暇名, 休暇種類）※シートが無ければスキップ
    - E) 特殊区分: special_attendance_type（シート名: 特殊区分 / 列: 区分コード, 区分名, 休日勤務, 有効）※シートが無ければスキップ
    """
    eng = _engine()
    xls = pd.ExcelFile(file)

    m = re.search(r"(\d+)班", file.name)
    team_name = f"{m.group(1)}班" if m else None

    with eng.begin() as con:
        if eng.url.get_backend_name() == "sqlite":
            con.exec_driver_sql("PRAGMA journal_mode=WAL;")
            con.exec_driver_sql("PRAGMA busy_timeout=5000;")

        # ---------- C) 社員 ----------
        if "社員" in xls.sheet_names:
            df_emp = pd.read_excel(xls, sheet_name="社員").rename(
                columns={
                    "社員番号": "employee_code",
                    "氏名": "name",
                    "部": "department_name",
                    "班": "team_name",
                    "社員タイプ": "employment_type",
                    "役職": "position",
                    "班長": "is_leader",
                    "副班長": "is_vice_leader",
                    "認証司": "is_certifier",
                    "勤務時間(日)": "default_work_hours",
                    "勤務時間(月)": "monthly_work_hours",
                }
            )

            for c in ["department_name", "team_name"]:
                if c in df_emp.columns:
                    df_emp[c] = df_emp[c].fillna("").astype(str).str.strip()

            dep_names = sorted(
                {n for n in df_emp.get("department_name", []).tolist() if n}
            )
            team_pairs = sorted(
                {
                    (r["department_name"], r["team_name"])
                    for _, r in df_emp[["department_name", "team_name"]].iterrows()
                    if str(r.get("team_name", "")).strip()
                }
            )

            for dn in dep_names:
                dep_id = con.execute(
                    text(
                        "SELECT department_id FROM department WHERE department_name=:dn"
                    ),
                    {"dn": dn},
                ).scalar_one_or_none()
                if dep_id is None:
                    con.execute(
                        text("INSERT INTO department (department_name) VALUES (:dn)"),
                        {"dn": dn},
                    )

            for dn, tn in team_pairs:
                dep_id = None
                if dn:
                    dep_id = con.execute(
                        text(
                            "SELECT department_id FROM department WHERE department_name=:dn"
                        ),
                        {"dn": dn},
                    ).scalar_one_or_none()
                    if dep_id is None:
                        con.execute(
                            text(
                                "INSERT INTO department (department_name) VALUES (:dn)"
                            ),
                            {"dn": dn},
                        )
                        dep_id = con.execute(
                            text(
                                "SELECT department_id FROM department WHERE department_name=:dn"
                            ),
                            {"dn": dn},
                        ).scalar_one()
                team_id = con.execute(
                    text("SELECT team_id FROM team WHERE team_name=:tn"), {"tn": tn}
                ).scalar_one_or_none()
                if team_id is None:
                    con.execute(
                        text(
                            "INSERT INTO team (team_name, department_id) VALUES (:tn, :did)"
                        ),
                        {"tn": tn, "did": dep_id},
                    )

            teams = pd.read_sql(text("SELECT team_id, team_name FROM team"), con)
            tmap = teams.set_index("team_name")["team_id"].to_dict()

            def to_bool(x):
                s = str(x).strip()
                return s in (
                    "True",
                    "TRUE",
                    "true",
                    "1",
                    "○",
                    "◯",
                    "Yes",
                    "YES",
                    "はい",
                )

            for c in ["is_leader", "is_vice_leader", "is_certifier"]:
                if c in df_emp.columns:
                    df_emp[c] = df_emp[c].apply(to_bool)

            for c in ["default_work_hours", "monthly_work_hours"]:
                if c in df_emp.columns:
                    df_emp[c] = pd.to_numeric(df_emp[c], errors="coerce")

            for _, r in df_emp.iterrows():
                code = str(r.get("employee_code") or "").strip()
                if not code:
                    continue
                team_id = None
                if "team_name" in r and str(r["team_name"]).strip():
                    team_id = tmap.get(str(r["team_name"]).strip())

                payload = {
                    "code": code,
                    "name": r.get("name"),
                    "employment_type": r.get("employment_type"),
                    "position": r.get("position"),
                    "is_leader": bool(r.get("is_leader", False)),
                    "is_vice_leader": bool(r.get("is_vice_leader", False)),
                    "is_certifier": bool(r.get("is_certifier", False)),
                    "dwh": None
                    if pd.isna(r.get("default_work_hours"))
                    else float(r.get("default_work_hours")),
                    "mwh": None
                    if pd.isna(r.get("monthly_work_hours"))
                    else float(r.get("monthly_work_hours")),
                    "team_id": team_id,
                }

                con.execute(
                    text(
                        """
                    INSERT INTO employee (
                        employee_code, name, employment_type, position,
                        is_leader, is_vice_leader, is_certifier,
                        default_work_hours, monthly_work_hours,
                        team_id, updated_at
                    )
                    VALUES (
                        :code, :name, :employment_type, :position,
                        :is_leader, :is_vice_leader, :is_certifier,
                        :dwh, :mwh, :team_id, CURRENT_TIMESTAMP
                    )
                    ON CONFLICT (employee_code) DO UPDATE SET
                      name               = EXCLUDED.name,
                      employment_type    = EXCLUDED.employment_type,
                      position           = EXCLUDED.position,
                      is_leader          = EXCLUDED.is_leader,
                      is_vice_leader     = EXCLUDED.is_vice_leader,
                      is_certifier       = EXCLUDED.is_certifier,
                      default_work_hours = EXCLUDED.default_work_hours,
                      monthly_work_hours = EXCLUDED.monthly_work_hours,
                      team_id            = COALESCE(EXCLUDED.team_id, employee.team_id),
                      updated_at         = CURRENT_TIMESTAMP
                """
                    ),
                    payload,
                )

        # ---------- B) 区情報 ----------
        if "区情報" in xls.sheet_names:
            dfz = pd.read_excel(xls, sheet_name="区情報").rename(
                columns={
                    "区コード": "zone_code",
                    "区名": "zone_name",
                    "班": "team_name",
                    "稼働": "operational_status",
                    "月": "mon",
                    "火": "tue",
                    "水": "wed",
                    "木": "thu",
                    "金": "fri",
                    "土": "sat",
                    "日": "sun",
                    "祝": "holiday",
                    "シフトタイプ": "shift_type",
                }
            )
            for c in [
                "team_name",
                "zone_code",
                "zone_name",
                "operational_status",
                "shift_type",
            ]:
                if c in dfz.columns:
                    dfz[c] = dfz[c].fillna("").astype(str).str.strip()

            for tn in sorted({t for t in dfz.get("team_name", []).tolist() if t}):
                tid = con.execute(
                    text("SELECT team_id FROM team WHERE team_name=:tn"), {"tn": tn}
                ).scalar_one_or_none()
                if tid is None:
                    con.execute(
                        text("INSERT INTO team (team_name) VALUES (:tn)"), {"tn": tn}
                    )

            for _, r in dfz.iterrows():
                tn = r.get("team_name")
                zc = r.get("zone_code")
                zn = r.get("zone_name")
                op = r.get("operational_status")
                st = r.get("shift_type")
                if not tn:
                    continue
                tid = con.execute(
                    text("SELECT team_id FROM team WHERE team_name=:tn"), {"tn": tn}
                ).scalar_one_or_none()
                if not tid:
                    continue
                if zc:
                    zid = con.execute(
                        text("SELECT zone_id FROM zone WHERE zone_code=:zc"), {"zc": zc}
                    ).scalar_one_or_none()
                    if zid is None:
                        con.execute(
                            text(
                                """
                            INSERT INTO zone (team_id, zone_code, zone_name, operational_status, shift_type)
                            VALUES (:tid,:zc,:zn,:op,:st)
                        """
                            ),
                            {"tid": tid, "zc": zc, "zn": zn, "op": op, "st": st},
                        )
                else:
                    zid = con.execute(
                        text(
                            """
                        SELECT z.zone_id FROM zone z
                         JOIN team t ON t.team_id = z.team_id
                        WHERE t.team_name = :tn AND z.zone_name = :zn
                    """
                        ),
                        {"tn": tn, "zn": zn},
                    ).scalar_one_or_none()
                    if zid is None:
                        con.execute(
                            text(
                                """
                            INSERT INTO zone (team_id, zone_name, operational_status, shift_type)
                        VALUES (:tid,:zn,:op,:st)
                        """
                            ),
                            {"tid": tid, "zn": zn, "op": op, "st": st},
                        )

            zones2 = pd.read_sql(
                text("SELECT zone_id, zone_code, zone_name, team_id FROM zone"), con
            )
            zmap_code = zones2.set_index("zone_code")["zone_id"].dropna().to_dict()
            zones2["tn"] = (
                zones2["team_id"].astype(str) + "||" + zones2["zone_name"].fillna("")
            )
            zmap_name = zones2.set_index("tn")["zone_id"].to_dict()

            for _, r in dfz.iterrows():
                tn = str(r.get("team_name") or "").strip()
                zc = str(r.get("zone_code") or "").strip()
                zn = str(r.get("zone_name") or "").strip()

                zid = None
                if zc:
                    zid = zmap_code.get(zc)
                if zid is None and tn and zn:
                    tid = con.execute(
                        text("SELECT team_id FROM team WHERE team_name=:tn"), {"tn": tn}
                    ).scalar_one_or_none()
                    if tid:
                        key = f"{tid}||{zn}"
                        zid = zmap_name.get(key)
                if not zid:
                    continue

                payload = {
                    "z": int(zid),
                    "mon": int(r.get("mon", 0) or 0),
                    "tue": int(r.get("tue", 0) or 0),
                    "wed": int(r.get("wed", 0) or 0),
                    "thu": int(r.get("thu", 0) or 0),
                    "fri": int(r.get("fri", 0) or 0),
                    "sat": int(r.get("sat", 0) or 0),
                    "sun": int(r.get("sun", 0) or 0),
                    "hol": int(r.get("holiday", 0) or 0),
                }
                con.execute(
                    text(
                        """
                    INSERT INTO demandprofile
                      (zone_id, demand_mon, demand_tue, demand_wed, demand_thu,
                       demand_fri, demand_sat, demand_sun, demand_holiday)
                    VALUES (:z,:mon,:tue,:wed,:thu,:fri,:sat,:sun,:hol)
                    ON CONFLICT (zone_id) DO UPDATE SET
                      demand_mon=EXCLUDED.demand_mon,
                      demand_tue=EXCLUDED.demand_tue,
                      demand_wed=EXCLUDED.demand_wed,
                      demand_thu=EXCLUDED.demand_thu,
                      demand_fri=EXCLUDED.demand_fri,
                      demand_sat=EXCLUDED.demand_sat,
                      demand_sun=EXCLUDED.demand_sun,
                      demand_holiday=EXCLUDED.demand_holiday
                """
                    ),
                    payload,
                )

        # ---------- A) 社員別需要 ----------
        if "社員別需要" in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name="社員別需要").rename(
                columns={
                    "社員番号": "employee_code",
                    "氏名": "name",
                    "月": "mon",
                    "火": "tue",
                    "水": "wed",
                    "木": "thu",
                    "金": "fri",
                    "土": "sat",
                    "日": "sun",
                    "祝": "hol",
                }
            )
            fixed = {
                "employee_code",
                "name",
                "mon",
                "tue",
                "wed",
                "thu",
                "fri",
                "sat",
                "sun",
                "hol",
            }
            zone_name_cols = [c for c in df.columns if c not in fixed]

            emp = pd.read_sql(
                text("SELECT employee_id, employee_code FROM employee"), con
            )
            emap = emp.set_index("employee_code")["employee_id"].astype(int).to_dict()

            zones = pd.read_sql(
                text(
                    """
                SELECT z.zone_id, z.zone_name, t.team_name
                  FROM zone z JOIN team t ON z.team_id=t.team_id
            """
                ),
                con,
            )
            zmap = (
                zones.set_index(["team_name", "zone_name"])["zone_id"]
                .astype(int)
                .to_dict()
            )

            if all(
                c in df.columns
                for c in ["mon", "tue", "wed", "thu", "fri", "sat", "sun", "hol"]
            ):
                for _, r in df.iterrows():
                    eid = emap.get(str(r.get("employee_code")))
                    if eid is None:
                        continue
                    vals = {
                        c: bool(r.get(c))
                        for c in [
                            "mon",
                            "tue",
                            "wed",
                            "thu",
                            "fri",
                            "sat",
                            "sun",
                            "hol",
                        ]
                    }
                    con.execute(
                        text(
                            """
                        INSERT INTO employee_availabilities
                          (employee_id, available_mon, available_tue, available_wed, available_thu,
                           available_fri, available_sat, available_sun, available_hol)
                        VALUES (:eid,:mon,:tue,:wed,:thu,:fri,:sat,:sun,:hol)
                        ON CONFLICT (employee_id) DO UPDATE SET
                          available_mon=EXCLUDED.available_mon,
                          available_tue=EXCLUDED.available_tue,
                          available_wed=EXCLUDED.available_wed,
                          available_thu=EXCLUDED.available_thu,
                          available_fri=EXCLUDED.available_fri,
                          available_sat=EXCLUDED.available_sat,
                          available_sun=EXCLUDED.available_sun,
                          available_hol=EXCLUDED.available_hol
                    """
                        ),
                        {"eid": int(eid), **vals},
                    )

            for _, r in df.iterrows():
                eid = emap.get(str(r.get("employee_code")))
                if eid is None:
                    continue
                for zn in zone_name_cols:
                    v = r.get(zn)
                    if pd.isna(v):
                        continue
                    zid = zmap.get((team_name, str(zn))) if team_name else None
                    if zid is None:
                        continue
                    con.execute(
                        text(
                            """
                        INSERT INTO employeezoneproficiency (employee_id, zone_id, proficiency)
                        VALUES (:eid,:zid,:p)
                        ON CONFLICT (employee_id, zone_id) DO UPDATE SET proficiency=EXCLUDED.proficiency
                    """
                        ),
                        {"eid": int(eid), "zid": int(zid), "p": int(v)},
                    )

        # ---------- D) 休暇種類 ----------
        if _has_table(eng, "leavetype") and "休暇種類" in xls.sheet_names:
            df_raw = pd.read_excel(xls, sheet_name="休暇種類")
            expected = ["休暇コード", "休暇名", "休暇種類"]
            missing = [c for c in expected if c not in df_raw.columns]
            if missing:
                raise ValueError(
                    f"休暇種類シートに必須列がありません: {missing}（想定: {expected}）"
                )

            df_lt = df_raw.rename(
                columns={
                    "休暇コード": "leave_code",
                    "休暇名": "leave_name",
                    "休暇種類": "leave_category",
                }
            )[["leave_code", "leave_name", "leave_category"]].copy()

            for c in ["leave_code", "leave_name", "leave_category"]:
                df_lt[c] = df_lt[c].astype(str).str.strip()

            df_lt = df_lt[df_lt["leave_code"] != ""]

            for _, r in df_lt.iterrows():
                con.execute(
                    text(
                        """
                    INSERT INTO leavetype (leave_code, leave_name, leave_category, updated_at)
                    VALUES (:code, :name, :cat, CURRENT_TIMESTAMP)
                    ON CONFLICT (leave_code) DO UPDATE SET
                      leave_name     = EXCLUDED.leave_name,
                      leave_category = EXCLUDED.leave_category,
                      updated_at     = CURRENT_TIMESTAMP
                """
                    ),
                    {
                        "code": r.leave_code,
                        "name": r.leave_name,
                        "cat": r.leave_category,
                    },
                )

        # ---------- E) 特殊区分 ----------
        if _has_table(eng, "special_attendance_type") and "特殊区分" in xls.sheet_names:
            df_raw = pd.read_excel(xls, sheet_name="特殊区分")
            expected = ["区分コード", "区分名", "休日勤務", "有効"]
            missing = [c for c in expected if c not in df_raw.columns]
            if missing:
                raise ValueError(
                    f"特殊区分シートに必須列がありません: {missing}（想定: {expected}）"
                )

            df_sat = df_raw.rename(
                columns={
                    "区分コード": "attendance_code",
                    "区分名": "attendance_name",
                    "休日勤務": "holiday_work_flag",
                    "有効": "is_active",
                }
            )[
                ["attendance_code", "attendance_name", "holiday_work_flag", "is_active"]
            ].copy()

            for c in ["attendance_code", "attendance_name"]:
                df_sat[c] = df_sat[c].astype(str).str.strip()

            df_sat["holiday_work_flag"] = (
                df_sat["holiday_work_flag"].map(_truthy_int).fillna(0).astype(int)
            )
            df_sat["is_active"] = (
                df_sat["is_active"].map(_truthy_int).fillna(1).astype(int)
            )

            df_sat = df_sat[df_sat["attendance_code"] != ""]

            for _, r in df_sat.iterrows():
                con.execute(
                    text(
                        """
                    INSERT INTO special_attendance_type
                      (attendance_code, attendance_name, holiday_work_flag, is_active, updated_at)
                    VALUES (:code, :name, :hol, :act, CURRENT_TIMESTAMP)
                    ON CONFLICT (attendance_code) DO UPDATE SET
                      attendance_name   = EXCLUDED.attendance_name,
                      holiday_work_flag = EXCLUDED.holiday_work_flag,
                      is_active         = EXCLUDED.is_active,
                      updated_at        = CURRENT_TIMESTAMP
                """
                    ),
                    {
                        "code": str(r.attendance_code),
                        "name": str(r.attendance_name),
                        "hol": int(r.holiday_work_flag),
                        "act": int(r.is_active),
                    },
                )

    typer.echo(f"✅ Imported to DB from: {file}")


if __name__ == "__main__":
    app()
