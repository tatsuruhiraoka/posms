"""
posms.optimization.shift_builder
================================
Excel から呼び出して「分担表案」だけを作る（PuLPで最適化）。
仕上げ（案→分担表／勤務指定表）は別マクロ側で実施。

【Excel前提】
- 予測: シート '予測' に '日付'(必須), '通常郵便'(任意), 'is_weekend_or_holiday'(任意)
- 班（スキル表）: 既定は最初のシート or team_sheet='班'
  * 列: '氏名' とジョブ列（0/1 等で担当可。0=事前指定のみ）
- 分担表案: A1 にヘッダ (shift_date | emp_id | 備考), A2 から本体を書き込む
"""

from __future__ import annotations

import os
import logging
from pathlib import Path
from typing import Dict, Tuple, Optional, List, Iterable

import pandas as pd
import pulp
import jpholiday
from datetime import date, datetime

# NEW: xlwingsを使わず書き込み。openpyxlで既存ブックに直接書く
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    # 任意: 互換のために残す（Excel から直接呼びたい場合のみ）
    import xlwings as xw  # type: ignore
except Exception:  # xlwings不在でも動く
    xw = None  # noqa

LOGGER = logging.getLogger("posms.optimization.shift_builder")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# ---------------- Excel 定数 ----------------
SHT_PRED = "予測"
SHT_DRAFT = "分担表案"

# ---------------- モデル用マッピング ----------------
SHIFT_TYPES = ['早番', '日勤', '夜勤', '通し']
REST_TYPES = [
    '週休', '非番', '祝休','計年', '年休', '夏休', '冬休', '代休',
    '承欠', '休職', '産休', '育休', '介護', '病休', 'その他'
]
SPECIAL_ATT = ['廃休', 'マル超']

DEFAULT_REQ_MAPPING = {
    "Weekday": {
        '早番': {'1区','4区','大口1','大口1-3'},
        '日勤': {'2区','3区','組立','補助'},
        '夜勤': {'速夜'},
        '通し': set(),
    },
    "Saturday": {
        '早番': {'速早'},
        '日勤': {'土曜補助'},
        '夜勤': set(),
        '通し': {'通し'},
    },
    "Sunday": {
        '早番': {'速早'},
        '日勤': set(),
        '夜勤': set(),
        '通し': {'通し'},
    },
    "Holiday": {
        '早番': {'速早'},
        '日勤': set(),
        '夜勤': set(),
        '通し': {'通し'},
    },
}

def get_day_type(d: date) -> str:
    wd = d.weekday()
    if wd == 5:
        return 'Saturday'
    elif wd == 6:
        return 'Sunday'
    elif jpholiday.is_holiday(d):
        return 'Holiday'
    else:
        return 'Weekday'


class ShiftBuilder:
    def __init__(
        self,
        excel_file_path: str,
        team_sheet: Optional[str] = None,
        req_mapping: Optional[Dict[str, Dict[str, set]]] = None,
    ):
        self.excel_file_path = excel_file_path
        self.team_sheet = team_sheet
        self.req_mapping = req_mapping or DEFAULT_REQ_MAPPING

        self.pred_df: pd.DataFrame = pd.DataFrame()
        self.df: pd.DataFrame = pd.DataFrame()

        self.days: List[date] = []
        self.week: List[List[date]] = []
        self.employees: List[str] = []
        self.shift_types: List[str] = SHIFT_TYPES
        self.jobs: List[str] = []

        self.saturday: List[date] = []
        self.sunday: List[date] = []
        self.holiday: List[date] = []
        self.weekday: List[date] = []

        self.assignable_jobs: Dict[str, Dict[str, float]] = {}
        self.pre_assigned_jobs: Dict[str, Dict[str, float]] = {}

        self.req: Dict[Tuple[date, str, str], int] = {}

        self.devPos: Dict[Tuple[str, int], pulp.LpVariable] = {}
        self.devNeg: Dict[Tuple[str, int], pulp.LpVariable] = {}
        self.priority_map: Dict[Tuple[date, str], float] = {}

        self.x = None
        self.y = None
        self.x_special = None
        self.rest_vars = None
        self.missing = None
        self.normal_work = None

        self.pre_dict_work: Dict[Tuple[str, date, str, str], int] = {}
        self.pre_dict_rest: Dict[Tuple[str, date], str] = {}
        self.pre_dict_special: Dict[Tuple[str, date], str] = {}

        self.model = pulp.LpProblem("Shift_Scheduling", pulp.LpMinimize)

        self.load_excel_data()

    # ---------------- 読み込み＆前処理 ----------------
    def load_excel_data(self):
        try:
            pred = pd.read_excel(self.excel_file_path, sheet_name=SHT_PRED)
        except Exception as e:
            raise FileNotFoundError(f"予測シート '{SHT_PRED}' の読込に失敗: {e}")

        date_cols = [c for c in pred.columns if str(c).strip() in ["日付", "date", "Date", "年月日", "yyyymmdd"]]
        if not date_cols:
            raise KeyError("予測シートに '日付' 列（または date/年月日/yyyymmdd）が見つかりません。")
        pred = pred.rename(columns={date_cols[0]: "日付"})
        pred["日付"] = pd.to_datetime(pred["日付"])
        pred = pred.dropna(subset=["日付"]).sort_values("日付").set_index("日付")

        if "is_weekend_or_holiday" not in pred.columns:
            pred["is_weekend_or_holiday"] = pred.index.date.map(
                lambda d: 1 if (d.weekday() >= 5 or jpholiday.is_holiday(d)) else 0
            )

        self.pred_df = pred.copy()
        self.days = [d.date() for d in self.pred_df.index]
        if not self.days:
            raise ValueError("予測シートから日付が取得できませんでした。")

        xls = pd.ExcelFile(self.excel_file_path)
        sheet_to_use = self.team_sheet
        if sheet_to_use is None:
            other_sheets = [s for s in xls.sheet_names if s != SHT_PRED]
            if not other_sheets:
                raise ValueError("スキル表のシートが見つかりません。")
            sheet_to_use = other_sheets[0]

        self.df = pd.read_excel(self.excel_file_path, sheet_name=sheet_to_use, header=1)
        if "氏名" not in self.df.columns:
            raise KeyError(f"スキル表シート '{sheet_to_use}' に '氏名' 列がありません。")

        self.jobs = list(self.df.columns[1:16])
        if not self.jobs:
            raise ValueError("ジョブ列が見つかりません（'氏名' の右側に 1 列以上必要）。")

        self.employees = [str(x) for x in self.df["氏名"].dropna().astype(str).tolist()]

        cols_assign = list(self.df.columns[1:min(13, len(self.df.columns))])
        subset_df = self.df[["氏名"] + cols_assign].copy()
        self.assignable_jobs = (
            subset_df.set_index("氏名")
            .apply(lambda row: {k: v for k, v in row.items() if pd.to_numeric(v, errors="coerce") and float(v) > 0}, axis=1)
            .to_dict()
        )

        cols_pre = list(self.df.columns[1:min(16, len(self.df.columns))])
        subset_df2 = self.df[["氏名"] + cols_pre].copy()
        self.pre_assigned_jobs = (
            subset_df2.set_index("氏名")
            .apply(lambda row: {k: v for k, v in row.items() if pd.to_numeric(v, errors="coerce") == 0}, axis=1)
            .to_dict()
        )

        self.week = [self.days[i:i + 7] for i in range(0, len(self.days), 7)]

        self.saturday, self.sunday, self.holiday, self.weekday = [], [], [], []
        for d in self.days:
            dt = get_day_type(d)
            if dt == "Saturday":
                self.saturday.append(d)
            elif dt == "Sunday":
                self.sunday.append(d)
            elif dt == "Holiday":
                self.holiday.append(d)
            else:
                self.weekday.append(d)

        self.set_req()

    # ---------------- 決定変数 ----------------
    def define_variables(self):
        self.x = pulp.LpVariable.dicts(
            "Shift",
            [(i, d, t, k) for i in self.employees for d in self.days for t in self.shift_types for k in self.jobs],
            cat='Binary'
        )
        self.y = pulp.LpVariable.dicts(
            "WorkDay",
            [(i, d) for i in self.employees for d in self.days],
            cat='Binary'
        )
        self.x_special = pulp.LpVariable.dicts(
            "x_special",
            [(i, d, s) for i in self.employees for d in self.days for s in SPECIAL_ATT],
            cat='Binary'
        )
        self.rest_vars = pulp.LpVariable.dicts(
            "Rest",
            [(i, d, r) for i in self.employees for d in self.days for r in REST_TYPES],
            cat='Binary'
        )
        self.missing = pulp.LpVariable.dicts(
            "Missing",
            [(d, t, k) for (d, t, k) in self.req.keys()],
            cat='Binary'
        )
        self.normal_work = pulp.LpVariable.dicts(
            "NormalWork",
            [(i, d) for i in self.employees for d in self.days],
            cat='Binary'
        )

    # ---------------- 需要設定 ----------------
    def set_req(self):
        self.req = {}
        jobs_set = set(self.jobs)
        for d in self.days:
            dt = get_day_type(d)
            if dt not in self.req_mapping:
                continue
            for t in self.shift_types:
                areas: Iterable[str] = self.req_mapping[dt].get(t, set())
                for k in areas:
                    if k in jobs_set:
                        self.req[(d, t, k)] = 1
                    else:
                        LOGGER.warning("需要ジョブ '%s' が班シートの列に存在しないためスキップ（%s, %s）。", k, d, t)

    # ---------------- 制約 ----------------
    def add_constraints(self):
        self.add_rest_constraints()
        self.add_special_constraints()
        self.add_work_constraints()
        self.add_work_or_rest_constraints()
        self.add_one_shift_per_workday()
        self.add_assignability_constraints()
        self.add_holiday_constraints()
        self.add_continuous_work_constraints()
        self.calculate_priority_map()
        self.setup_coverage_and_objective()

    def add_rest_constraints(self):
        for i in self.employees:
            for d in self.days:
                if (i, d) in self.pre_dict_rest:
                    r_type = self.pre_dict_rest[(i, d)]
                    self.model += self.rest_vars[(i, d, r_type)] == 1, f"FixRest_{i}_{d}_{r_type}"
                    for r in REST_TYPES:
                        if r != r_type:
                            self.model += self.rest_vars[(i, d, r)] == 0, f"ZeroRest_{i}_{d}_{r}"
                    for t in self.shift_types:
                        for k in self.jobs:
                            self.model += self.x[(i, d, t, k)] == 0, f"ZeroWork_{i}_{d}_{t}_{k}"
                    for s in SPECIAL_ATT:
                        self.model += self.x_special[(i, d, s)] == 0, f"NoSpecial_{i}_{d}_{s}"
                else:
                    for r in REST_TYPES:
                        if r not in ["週休", "非番", "祝休"]:
                            self.model += self.rest_vars[(i, d, r)] == 0, f"ZeroRest_{i}_{d}_{r}"

    def add_special_constraints(self):
        for i in self.employees:
            for d in self.days:
                if (i, d) in self.pre_dict_special:
                    s_type = self.pre_dict_special[(i, d)]
                    for s in SPECIAL_ATT:
                        self.model += self.x_special[(i, d, s)] == (1 if s == s_type else 0), f"Spec_{i}_{d}_{s}"
                else:
                    for s in SPECIAL_ATT:
                        self.model += self.x_special[(i, d, s)] == 0, f"NoSpec_{i}_{d}_{s}"

    def add_work_constraints(self):
        for i in self.employees:
            for d in self.days:
                for t in self.shift_types:
                    for k in self.jobs:
                        if (i, d, t, k) in self.pre_dict_work:
                            self.model += self.x[(i, d, t, k)] == 1, f"ForceWork_{i}_{d}_{t}_{k}"
                        else:
                            if (d, t, k) not in self.req:
                                self.model += self.x[(i, d, t, k)] == 0, f"NoReq_{i}_{d}_{t}_{k}"

    def add_work_or_rest_constraints(self):
        for i in self.employees:
            for d in self.days:
                self.model += (
                    self.y[(i, d)] + pulp.lpSum(self.rest_vars[(i, d, r)] for r in REST_TYPES)
                ) == 1, f"WorkOrRest_{i}_{d}"

    def add_one_shift_per_workday(self):
        for i in self.employees:
            for d in self.days:
                sum_x_id = pulp.lpSum(self.x[(i, d, t, k)] for t in self.shift_types for k in self.jobs)
                self.model += sum_x_id == self.y[(i, d)], f"OneShiftOrNone_{i}_{d}"

    def add_assignability_constraints(self):
        for i in self.employees:
            for d in self.days:
                for t in self.shift_types:
                    for k in self.jobs:
                        if k in self.assignable_jobs.get(i, {}):
                            if (i, d, t, k) in self.pre_dict_work:
                                self.model += self.x[(i, d, t, k)] == 1, f"MustWork_{i}_{d}_{t}_{k}"
                        elif k in self.pre_assigned_jobs.get(i, {}):
                            if (i, d, t, k) in self.pre_dict_work:
                                self.model += self.x[(i, d, t, k)] == 1, f"MustWork_{i}_{d}_{t}_{k}"
                            else:
                                self.model += self.x[(i, d, t, k)] == 0, f"NoWork_{i}_{d}_{t}_{k}"
                        else:
                            self.model += self.x[(i, d, t, k)] == 0, f"InvalidJob_{i}_{d}_{t}_{k}"

    def add_holiday_constraints(self):
        for i in self.employees:
            for d in self.days:
                self.model += self.rest_vars[(i, d, '週休')] + self.x_special[(i, d, '廃休')] <= 1
                self.model += self.rest_vars[(i, d, '非番')] + self.x_special[(i, d, 'マル超')] <= 1

        for i in self.employees:
            self.model += (
                pulp.lpSum(self.rest_vars[(i, d, '週休')] + self.x_special[(i, d, '廃休')] for d in self.days) == 4
            ), f"WeeklyOff_{i}"
            self.model += (
                pulp.lpSum(self.rest_vars[(i, d, '非番')] + self.x_special[(i, d, 'マル超')] for d in self.days) >= 4
            ), f"Hiban_{i}"

        for i in self.employees:
            for w, days_in_week in enumerate(self.week):
                self.model += (
                    pulp.lpSum(self.rest_vars[(i, d, '週休')] + self.x_special[(i, d, '廃休')] for d in days_in_week) == 1
                ), f"WeeklyOneOff_{i}_{w}"
                self.devPos[(i, w)] = pulp.LpVariable(f"devPos_{i}_{w}", lowBound=0)
                self.devNeg[(i, w)] = pulp.LpVariable(f"devNeg_{i}_{w}", lowBound=0)
                sumHibanOrMarucho = pulp.lpSum(
                    self.rest_vars[(i, d, '非番')] + self.x_special[(i, d, 'マル超')] for d in days_in_week
                )
                self.model += sumHibanOrMarucho - 1 <= self.devPos[(i, w)]
                self.model += 1 - sumHibanOrMarucho <= self.devNeg[(i, w)]

        for i in self.employees:
            for d in self.sunday:
                self.model += self.y[(i, d)] + self.rest_vars[(i, d, '週休')] == 1, f"SundayWorkOrOff_{i}_{d}"

        for i in self.employees:
            for d in self.saturday:
                if (i, d) in self.pre_dict_rest:
                    continue
                self.model += self.y[(i, d)] + self.rest_vars[(i, d, '非番')] == 1, f"SaturdayWorkOrHiban_{i}_{d}"

        for i in self.employees:
            for d in self.days:
                if d not in self.holiday:
                    self.model += self.rest_vars[(i, d, '祝休')] == 0, f"NoHolidayRest_{i}_{d}"

        for d in self.holiday:
            for i in self.employees:
                if (i, d) in self.pre_dict_rest and self.pre_dict_rest[(i, d)] == '非番':
                    continue
                self.model += self.y[(i, d)] + self.rest_vars[(i, d, '祝休')] == 1, f"HolidayWorkOrHolidayRest_{i}_{d}"

    def add_continuous_work_constraints(self):
        for i in self.employees:
            for d_idx in range(len(self.days) - 10 + 1):
                consecutive = [self.days[d_idx + off] for off in range(10)]
                self.model += pulp.lpSum(self.y[(i, dd)] for dd in consecutive) <= 9, f"No10Consecutive_{i}_{d_idx}"

        for i in self.employees:
            for d in self.days:
                self.model += self.normal_work[(i, d)] >= self.y[(i, d)] - self.x_special[(i, d, '廃休')] - self.x_special[(i, d, 'マル超')]
                self.model += self.normal_work[(i, d)] <= 1 - self.x_special[(i, d, '廃休')]
                self.model += self.normal_work[(i, d)] <= 1 - self.x_special[(i, d, 'マル超')]

        for i in self.employees:
            for start_idx in range(len(self.days) - 6 + 1):
                consecutive = [self.days[start_idx + r] for r in range(6)]
                self.model += pulp.lpSum(self.normal_work[(i, d_)] for d_ in consecutive) <= 5, f"No6Normal_{i}_{start_idx}"

    def calculate_supply_and_demand(self):
        employee_possible_days = {}
        for i in self.employees:
            base = 20  # 28日 - 週休8（目安）
            extra_rest = 0
            for d in self.days:
                if (i, d) in self.pre_dict_rest and self.pre_dict_rest[(i, d)] in [
                    '計年','年休','夏休','冬休','代休','承欠','休職','産休','育休','介護','病休','その他'
                ]:
                    extra_rest += 1
            employee_possible_days[i] = base - extra_rest
            for d in self.days:
                if (i, d) in self.pre_dict_special and self.pre_dict_special[(i, d)] in ['廃休','マル超']:
                    employee_possible_days[i] += 1

        num_holiday = len(self.holiday)
        total_supply = sum(employee_possible_days.values())
        total_supply_adj = total_supply - (len(self.employees) - 2) * num_holiday
        base_demand = 200 - 20
        total_demand = base_demand - (7 * num_holiday)
        excess = total_supply_adj - total_demand

        LOGGER.info("祝日=%d, demand(base)=%d, supply=%d (adj=%d), 過不足=%d",
                    num_holiday, total_demand, total_supply, total_supply_adj, excess)
        return excess

    def calculate_priority_map(self):
        excess_or_def = self.calculate_supply_and_demand()

        base_priority = {
            '1区':10, '2区':10, '3区':10, '4区':10,
            '大口1':10, '大口1-3':10,
            '組立':1, '補助':1,
            '速夜':5, '土曜補助':10,
            '速早':10, '通し':10
        }
        for d in self.days:
            for k, val in base_priority.items():
                self.priority_map[(d, k)] = val

        if "通常郵便" in self.pred_df.columns:
            if excess_or_def > 0:
                top_days = self.pred_df.nlargest(min(excess_or_def, len(self.pred_df)), '通常郵便')
                for ts in top_days.index:
                    dd = ts.date()
                    if (dd, '補助') in self.priority_map:
                        self.priority_map[(dd, '補助')] += 1
            elif excess_or_def < 0:
                if "is_weekend_or_holiday" in self.pred_df.columns:
                    df_wd = self.pred_df[self.pred_df["is_weekend_or_holiday"] == 0]
                else:
                    df_wd = self.pred_df[[ts.weekday() < 5 for ts in self.pred_df.index]]
                need = min(abs(excess_or_def), len(df_wd))
                for ts in df_wd.nsmallest(need, '通常郵便').index:
                    dd = ts.date()
                    if (dd, '速夜') in self.priority_map:
                        self.priority_map[(dd, '速夜')] = max(0, self.priority_map[(dd, '速夜')] - 2)

    def setup_coverage_and_objective(self):
        for (d, t, k), needed in self.req.items():
            self.model += (
                pulp.lpSum(self.x[(i, d, t, k)] for i in self.employees)
                == needed * (1 - self.missing[(d, t, k)])
            ), f"Coverage_{d}_{t}_{k}"

        obj_missing = pulp.lpSum(
            self.priority_map.get((d, k), 0) * self.missing[(d, t, k)]
            for (d, t, k) in self.req.keys()
        )
        obj_weekdev = pulp.lpSum(
            self.devPos[(i, w)] + self.devNeg[(i, w)]
            for i in self.employees
            for w, _days in enumerate(self.week)
        )
        obj_total_work = 0.001 * pulp.lpSum(self.y[(i, d)] for i in self.employees for d in self.days)
        self.model += obj_missing + obj_weekdev + obj_total_work

    # ---------------- 解く ----------------
    def solve(self, solver: Optional[pulp.LpSolver] = None) -> str:
        if solver is None:
            solver = pulp.PULP_CBC_CMD(msg=False)
        result = self.model.solve(solver)
        status_str = pulp.LpStatus[self.model.status]
        LOGGER.info("Solve status: %s", status_str)
        return status_str

    # ---------------- “案” の抽出 ----------------
    def _collect_draft_rows(self) -> List[Dict[str, object]]:
        rows = []
        for d in self.days:
            for i in self.employees:
                yv = self.y[(i, d)].varValue if self.y[(i, d)] is not None else 0
                s1 = self.x_special[(i, d, '廃休')].varValue if self.x_special[(i, d, '廃休')] is not None else 0
                s2 = self.x_special[(i, d, 'マル超')].varValue if self.x_special[(i, d, 'マル超')] is not None else 0
                if (yv == 1) or (s1 == 1) or (s2 == 1):
                    rows.append({"shift_date": d, "emp_id": i, "備考": "案"})
        if rows:
            df = pd.DataFrame(rows).drop_duplicates().sort_values(["shift_date", "emp_id"])
            return df.to_dict(orient="records")
        return rows

    # ---------------- Excel書き込み（openpyxl） ----------------
    def write_draft_to_excel_path(self, excel_path: Path) -> Path:
        """
        ブック excel_path の '分担表案' シートの A2 から values を書込み。
        A1 はヘッダ（shift_date | emp_id | 備考）想定。下はクリアしてから書く。
        .xlsm の場合は keep_vba=True でマクロ温存。
        """
        rows = self._collect_draft_rows()
        wb = load_workbook(excel_path, keep_vba=excel_path.suffix.lower() == ".xlsm")
        if SHT_DRAFT not in wb.sheetnames:
            raise KeyError(f"シート '{SHT_DRAFT}' が見つかりません。")
        ws = wb[SHT_DRAFT]

        # 既存内容（A2 以降）クリア
        max_row = ws.max_row if ws.max_row else 2
        max_col = ws.max_column if ws.max_column else 3
        if max_row >= 2:
            ws.delete_rows(2, max_row - 1)

        if rows:
            # A1 がヘッダ前提、A2 から本体
            # 日付は文字列化（テンプレ表示形式を使う場合はこのままでOK）
            for r_idx, rec in enumerate(rows, start=2):
                ws[f"A{r_idx}"].value = pd.to_datetime(rec["shift_date"]).date().isoformat()
                ws[f"B{r_idx}"].value = rec["emp_id"]
                ws[f"C{r_idx}"].value = rec["備考"]

        wb.save(excel_path)
        return excel_path

    # ---------------- ワンショット（単体テスト用） ----------------
    def build_draft(self) -> pd.DataFrame:
        self.define_variables()
        self.add_constraints()
        self.solve()
        return pd.DataFrame(self._collect_draft_rows())


# ========= 互換: xlwings マクロ（残すが任意） =========
if xw is not None:
    @xw.sub
    def 作成_分担表案():
        wb = xw.Book.caller()
        excel_path = Path(wb.fullname)
        if not excel_path:
            raise RuntimeError("このマクロは保存済みのブックから実行してください。")
        builder = ShiftBuilder(excel_file_path=str(excel_path))
        builder.define_variables()
        builder.add_constraints()
        status = builder.solve()
        if status != "Optimal":
            raise RuntimeError(f"最適化に失敗しました（{status}）")
        builder.write_draft_to_excel_path(excel_path)


# ========= CLI エントリ =========
# 既存 Typer CLI に統合する場合は、以下を posms/cli.py 側に移して
# app.add_typer(...) などでぶら下げてもOK
import typer
cli_app = typer.Typer(add_completion=False)

@cli_app.command("shift-draft")
def cli_shift_draft(
    excel: Path = typer.Argument(..., help="予測/班/分担表案 を含むExcelファイル（.xlsm可）"),
    team_sheet: Optional[str] = typer.Option(None, help="班のシート名（未指定なら先頭の予測以外シート）"),
):
    sb = ShiftBuilder(str(excel), team_sheet=team_sheet)
    sb.define_variables()
    sb.add_constraints()
    status = sb.solve()
    if status != "Optimal":
        raise typer.Exit(code=2)
    sb.write_draft_to_excel_path(excel)
    typer.echo(f"分担表案を書き込みました: {excel}")

if __name__ == "__main__":
    typer.run(cli_shift_draft)
