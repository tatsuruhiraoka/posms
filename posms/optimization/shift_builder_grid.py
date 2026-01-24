# posms/optimization/shift_builder_grid.py
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Tuple, Optional, List
import datetime as dt

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


@dataclass(frozen=True)
class GridLayout:
    sheet_name: str = "分担予定表(案)"
    header_row: int = 22  # 日付が入っている行（テンプレに合わせて調整）
    start_row: int = 23  # 1人目の上段行
    name_col: int = 2  # B列=氏名（上段）
    first_day_col: int = 3  # C列=1日目
    last_day_col: int = 30  # AD列=28日目
    empno_col: Optional[int] = 31  # AE列=社員番号（あれば。無ければ None に）


def _cell_date_value(c: Cell) -> Optional[dt.date]:
    """Excelセルが date/datetime/文字列(YYYY-MM-DD など) のいずれでも date に寄せる"""
    v = c.value
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        s = v.strip()
        # 例: "2026-01-15" / "2026/01/15"
        s = s.replace("/", "-")
        try:
            return dt.date.fromisoformat(s)
        except Exception:
            return None
    return None


def _build_date_to_col(ws, layout: GridLayout) -> Dict[dt.date, int]:
    m: Dict[dt.date, int] = {}
    for col in range(layout.first_day_col, layout.last_day_col + 1):
        d = _cell_date_value(ws.cell(row=layout.header_row, column=col))
        if d:
            m[d] = col
    return m


def _iter_employees(ws, layout: GridLayout) -> List[Tuple[int, str, str]]:
    """
    2行で1人：上段行だけを見る。
    戻り: [(upper_row, emp_key, name), ...]
      emp_key: 社員番号があればそれ、無ければ氏名文字列
    """
    out = []
    r = layout.start_row
    while True:
        name = ws.cell(row=r, column=layout.name_col).value
        if name is None or str(name).strip() == "":
            break  # 連続ブロック終端とみなす
        name_s = str(name).strip()

        if layout.empno_col is not None:
            empno = ws.cell(row=r, column=layout.empno_col).value
            emp_key = (
                str(empno).strip()
                if empno is not None and str(empno).strip() != ""
                else name_s
            )
        else:
            emp_key = name_s

        out.append((r, emp_key, name_s))
        r += 2  # 次の人（2行で1人）
    return out


def apply_assignments_to_grid_xlsm(
    in_xlsm: Path,
    out_xlsm: Path,
    assignments: Dict[Tuple[str, dt.date], Tuple[Optional[str], Optional[str]]],
    layout: Optional[GridLayout] = None,
    clear_before_write: bool = True,
) -> Path:
    """
    assignments:
      key: (emp_key, date)
      val: (upper_text, lower_text)

    upper_text: 上段に入れる（例: 早番/日勤/夜勤/通し）
    lower_text: 下段に入れる（例: 1区/組立/週休/廃休/マル超 など）
    """
    layout = layout or GridLayout()
    wb = load_workbook(in_xlsm, keep_vba=True)
    if layout.sheet_name not in wb.sheetnames:
        raise RuntimeError(f"シート '{layout.sheet_name}' が見つかりません: {in_xlsm}")
    ws = wb[layout.sheet_name]

    date_to_col = _build_date_to_col(ws, layout)
    employees = _iter_employees(ws, layout)

    # クリア（上段/下段の入力欄だけ）
    if clear_before_write:
        for upper_row, _emp_key, _ in employees:
            lower_row = upper_row + 1
            for col in range(layout.first_day_col, layout.last_day_col + 1):
                ws.cell(row=upper_row, column=col).value = None
                ws.cell(row=lower_row, column=col).value = None

    # 書き込み
    for upper_row, emp_key, _ in employees:
        lower_row = upper_row + 1
        for d, col in date_to_col.items():
            k = (emp_key, d)
            if k not in assignments:
                continue
            upper_text, lower_text = assignments[k]
            if upper_text is not None:
                ws.cell(row=upper_row, column=col).value = str(upper_text)
            if lower_text is not None:
                ws.cell(row=lower_row, column=col).value = str(lower_text)

    out_xlsm.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsm)
    return out_xlsm
