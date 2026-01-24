from __future__ import annotations
from pathlib import Path
from typing import Dict, Optional

import pandas as pd
from shutil import copy2
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils import get_column_letter


def write_dataframe_to_excel(
    df: pd.DataFrame,
    out_path: Path,
    sheet_name: str = "export",
    template_path: Optional[Path] = None,
    header_map: Optional[Dict[str, str]] = None,
    start_cell: str = "A1",
    append: bool = False,
    *,
    verbose: bool = False,
) -> None:
    """
    - .xlsm は keep_vba=True で読込/保存（VBA温存）
    - append=False（既定）: シートを“削除→保存→再読込→同名再作成”してから書き込み（重複ゼロ）
    - append=True         : 既存ヘッダ一致なら末尾に追記
    - start_cell は "AB5" 形式対応
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    _df = df.copy()
    if header_map:
        _df.rename(columns=header_map, inplace=True)

    def _to_cell(v):
        if pd.isna(v):
            return None
        if hasattr(v, "item"):
            try:
                return v.item()
            except Exception:
                pass
        return v

    def _load(path: Path):
        return load_workbook(path, keep_vba=(path.suffix.lower() == ".xlsm"))

    def _ensure_base_workbook() -> None:
        """出力先が無ければテンプレを物理コピー。無ければ真っさらを保存して作る。"""
        if out_path.exists():
            return
        if template_path and Path(template_path).exists():
            copy2(template_path, out_path)  # 物理コピー（最も壊れにくい）
        else:
            wb0 = Workbook()
            wb0.remove(wb0.active)
            wb0.save(out_path)

    # まずベースのファイルを用意
    _ensure_base_workbook()

    # いったん常に出力ファイルを開く（既存/新規どちらも）
    wb = _load(out_path)

    # ==== append=False のときは“確実に”中身を消す ====
    if not append and sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        if verbose:
            print(f"[overwrite] remove sheet '{sheet_name}' at index {idx}")
        wb.remove(wb[sheet_name])
        # ここで一度保存 → 再読込（内部の表/オートフィルタ/UsedRangeの残りをリセット）
        wb.save(out_path)
        wb = _load(out_path)
        ws = wb.create_sheet(title=sheet_name, index=idx)
    else:
        # append=True か、まだ存在しない場合
        ws = (
            wb[sheet_name]
            if sheet_name in wb.sheetnames
            else wb.create_sheet(title=sheet_name)
        )

    # ==== 書き込み座標 ====
    row0, col0 = coordinate_to_tuple(start_cell)  # "AB5" → (5, 28)

    if append:
        # 既存ヘッダと一致するか判定（一致しなければヘッダから書く）
        write_header = True
        if ws.max_row >= row0 and _df.shape[1] > 0:
            existing = [
                ws.cell(row=row0, column=col0 + j).value for j in range(_df.shape[1])
            ]
            if existing == list(_df.columns):
                write_header = False

        # 実質最終行（対象列のどれかに値があれば使用中とみなす）
        last = ws.max_row
        while last >= row0 and all(
            ws.cell(row=last, column=col0 + j).value in (None, "")
            for j in range(_df.shape[1])
        ):
            last -= 1
        start_row = (row0 if write_header else row0 + 1) if last < row0 else last + 1
        base = 0 if write_header else -1

        if verbose:
            print(f"[append] header_match={not write_header}, start_row={start_row}")

        # ヘッダ
        if write_header:
            for j, name in enumerate(_df.columns):
                ws.cell(row=row0, column=col0 + j, value=str(name))

        # 値
        for i, row in enumerate(_df.itertuples(index=False)):
            for j, v in enumerate(row):
                ws.cell(
                    row=start_row + base + 1 + i, column=col0 + j, value=_to_cell(v)
                )
    else:
        # 完全上書き（シート作り直し済み）
        # ヘッダ
        for j, name in enumerate(_df.columns):
            ws.cell(row=row0, column=col0 + j, value=str(name))
        # 値
        for i, row in enumerate(_df.itertuples(index=False)):
            for j, v in enumerate(row):
                ws.cell(row=row0 + 1 + i, column=col0 + j, value=_to_cell(v))

    # 列幅（A〜ZZ…対応）
    n_cols = _df.shape[1]
    for j in range(n_cols):
        col_letter = get_column_letter(col0 + j)
        header_len = len(str(_df.columns[j]))
        max_len = header_len
        for v in _df.iloc[:, j]:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    wb.save(out_path)
